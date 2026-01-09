import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def generate_mockup():
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts"
    
    # Columns
    headers = ["Date", "Vendor", "Description", "Category", "Amount", "Currency", "Notes"]
    ws.append(headers)
    
    # Dummy Data
    ws.append(["2024-01-15", "Apple Store", "MacBook Pro M3", "Office Equipment", 2499.00, "USD", "Work laptop"])
    ws.append(["2024-01-16", "Starbucks", "Coffee with Client", "Meals", 15.50, "CAD", "Project kickoff"])
    ws.append(["2024-01-17", "Amazon", "USB-C Cables", "Office Supplies", 45.99, "CAD", ""])
    
    # 1. Format Header (Blue BG, White Bold Text)
    header_fill = PatternFill(start_color="3366CC", end_color="3366CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    # 2. Freeze Pane
    ws.freeze_panes = "A2"
    
    # 3. Format Amount Column (Currency)
    for row in range(2, 5):
        ws[f"E{row}"].number_format = "$#,##0.00"
        
    # 4. Set Column Widths
    widths = [12, 25, 30, 20, 15, 10, 20]
    for i, width in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = width
        
    wb.save("Format_Preview.xlsx")
    print("✅ Format_Preview.xlsx generated.")

if __name__ == "__main__":
    generate_mockup()
