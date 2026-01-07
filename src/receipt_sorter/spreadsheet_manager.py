"""
Spreadsheet Manager Module
Creates and updates Excel spreadsheets for each currency
"""

import os
import logging
from typing import Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from . import config

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class SpreadsheetManager:
    """Manage Excel spreadsheets for receipt tracking"""

    def __init__(self):
        """Initialize spreadsheet manager"""
        self.output_base = config.OUTPUT_BASE_FOLDER
        self.columns = config.SPREADSHEET_COLUMNS

    def get_spreadsheet_path(self, currency: str) -> str:
        """
        Get path to currency-specific spreadsheet

        Args:
            currency: Currency code (e.g., CAD, USD)

        Returns:
            Path to Excel file
        """
        currency = currency.upper()
        currency_folder = os.path.join(self.output_base, currency)
        spreadsheet_path = os.path.join(currency_folder, f"{currency}_Receipts.xlsx")
        return spreadsheet_path

    def initialize_spreadsheet(self, currency: str) -> bool:
        """
        Create new spreadsheet if it doesn't exist

        Args:
            currency: Currency code

        Returns:
            True if successful, False otherwise
        """
        try:
            spreadsheet_path = self.get_spreadsheet_path(currency)

            # Check if file already exists
            if os.path.exists(spreadsheet_path):
                logger.debug(f"Spreadsheet already exists: {spreadsheet_path}")
                return True

            # Create new DataFrame with columns
            df = pd.DataFrame(columns=self.columns)

            # Write to Excel
            df.to_excel(spreadsheet_path, index=False, sheet_name='Receipts')

            # Apply formatting
            self._apply_formatting(spreadsheet_path, currency)

            logger.info(f"Created new spreadsheet: {spreadsheet_path}")
            return True

        except Exception as e:
            logger.error(f"Error initializing spreadsheet for {currency}: {str(e)}")
            return False

    def add_receipt_entry(self, receipt_data: Dict, category: str,
                         confidence: int, organized_filepath: str) -> bool:
        """
        Add receipt entry to appropriate spreadsheet

        Args:
            receipt_data: Extracted receipt data
            category: Tax category
            confidence: Categorization confidence score
            organized_filepath: Path to organized PDF file

        Returns:
            True if successful, False otherwise
        """
        try:
            currency = receipt_data.get('currency', 'UNKNOWN')

            # Initialize spreadsheet if needed
            self.initialize_spreadsheet(currency)

            spreadsheet_path = self.get_spreadsheet_path(currency)

            # Read existing spreadsheet
            df = pd.read_excel(spreadsheet_path, sheet_name='Receipts')

            # Prepare new row
            new_row = {
                'Date': receipt_data.get('date', 'UNKNOWN'),
                'Vendor': receipt_data.get('vendor', 'UNKNOWN'),
                'Description': receipt_data.get('description', ''),
                'Category': category,
                'Amount': receipt_data.get('total_amount', 0),
                'Currency': currency,
                'File Name': os.path.basename(organized_filepath),
                'Notes': f"Confidence: {confidence}%" if confidence < 100 else ""
            }

            # Append new row
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

            # Sort by date (most recent first)
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            df = df.sort_values('Date', ascending=False)

            # Write back to Excel
            df.to_excel(spreadsheet_path, index=False, sheet_name='Receipts')

            # Reapply formatting
            self._apply_formatting(spreadsheet_path, currency)

            logger.info(f"Added entry to {currency} spreadsheet: {new_row['Vendor']} - {new_row['Amount']}")
            return True

        except Exception as e:
            logger.error(f"Error adding receipt entry: {str(e)}")
            return False

    def _apply_formatting(self, spreadsheet_path: str, currency: str):
        """
        Apply formatting to Excel spreadsheet

        Args:
            spreadsheet_path: Path to Excel file
            currency: Currency code for formatting amounts
        """
        try:
            # Load workbook
            wb = load_workbook(spreadsheet_path)
            ws = wb.active

            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Column widths
            column_widths = {
                'A': 12,  # Date
                'B': 25,  # Vendor
                'C': 40,  # Description
                'D': 22,  # Category
                'E': 12,  # Amount
                'F': 10,  # Currency
                'G': 30,  # File Name
                'H': 25   # Notes
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Format amount column as currency
            currency_symbol = self._get_currency_symbol(currency)
            for row in range(2, ws.max_row + 1):
                amount_cell = ws[f'E{row}']
                if amount_cell.value:
                    amount_cell.number_format = f'{currency_symbol}#,##0.00'

            # Format date column
            for row in range(2, ws.max_row + 1):
                date_cell = ws[f'A{row}']
                if date_cell.value:
                    date_cell.number_format = 'YYYY-MM-DD'

            # Add totals row
            last_row = ws.max_row + 2
            ws[f'D{last_row}'] = "TOTAL:"
            ws[f'D{last_row}'].font = Font(bold=True)
            ws[f'E{last_row}'] = f'=SUM(E2:E{last_row-2})'
            ws[f'E{last_row}'].font = Font(bold=True)
            ws[f'E{last_row}'].number_format = f'{currency_symbol}#,##0.00'

            # Add category breakdown
            categories = config.TAX_CATEGORIES
            breakdown_start = last_row + 2

            ws[f'D{breakdown_start}'] = "Category Breakdown:"
            ws[f'D{breakdown_start}'].font = Font(bold=True, size=11)

            for idx, category in enumerate(categories, start=1):
                row_num = breakdown_start + idx
                ws[f'D{row_num}'] = category
                # SUMIF formula to sum amounts for this category
                ws[f'E{row_num}'] = f'=SUMIF(D2:D{last_row-2},D{row_num},E2:E{last_row-2})'
                ws[f'E{row_num}'].number_format = f'{currency_symbol}#,##0.00'

            # Save workbook
            wb.save(spreadsheet_path)
            logger.debug(f"Applied formatting to {spreadsheet_path}")

        except Exception as e:
            logger.error(f"Error applying formatting: {str(e)}")

    def _get_currency_symbol(self, currency: str) -> str:
        """
        Get currency symbol for formatting

        Args:
            currency: Currency code

        Returns:
            Currency symbol
        """
        symbols = {
            'CAD': '$',
            'USD': '$',
            'EUR': '€',
            'GBP': '£',
            'JPY': '¥',
            'AUD': 'A$',
            'CHF': 'CHF '
        }
        return symbols.get(currency.upper(), '$')

    def generate_summary_report(self) -> Dict:
        """
        Generate summary report across all currency spreadsheets

        Returns:
            Dictionary with summary statistics
        """
        summary = {
            'total_receipts': 0,
            'currencies': {},
            'categories': {}
        }

        try:
            # Check each currency folder
            for currency in config.SUPPORTED_CURRENCIES:
                spreadsheet_path = self.get_spreadsheet_path(currency)

                if os.path.exists(spreadsheet_path):
                    df = pd.read_excel(spreadsheet_path, sheet_name='Receipts')

                    if not df.empty:
                        count = len(df)
                        total = df['Amount'].sum()

                        summary['total_receipts'] += count
                        summary['currencies'][currency] = {
                            'count': count,
                            'total': total
                        }

                        # Category breakdown
                        for category in config.TAX_CATEGORIES:
                            cat_total = df[df['Category'] == category]['Amount'].sum()
                            if cat_total > 0:
                                if category not in summary['categories']:
                                    summary['categories'][category] = 0
                                summary['categories'][category] += cat_total

            return summary

        except Exception as e:
            logger.error(f"Error generating summary report: {str(e)}")
            return summary
